from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponse
import io, os
from employee.models import Employee
from openpyxl import load_workbook
from django.utils import timezone
from .models_tupot import TUPOTExportHistory
import random, string

def is_superuser_or_pot(user):
    return user.is_superuser or user.groups.filter(name='pot').exists()

@user_passes_test(is_superuser_or_pot)
def tu_pot(request):
    error = None
    export_filename = None
    export_time = None
    download_url = None
    tu_export_filename = None
    tu_export_time = None
    tu_download_url = None
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            from django.utils import timezone
            upload_time = timezone.localtime()
            # Save original uploaded file
            upload_filename = f"tu_pot_upload_{upload_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            upload_path = f"media/tu_pot_uploads/{upload_filename}"
            os.makedirs('media/tu_pot_uploads', exist_ok=True)
            with open(upload_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)
            # Process uploaded file
            wb = load_workbook(upload_path)
            ws = wb.active
            max_col = ws.max_column
            max_row = ws.max_row
            if max_col < 10:
                error = 'File must have at least 10 columns.'
            else:
                # 1. Membership file
                yes_emails = set(Employee.objects.filter(membership_type_by_admin__name__iexact='Yes').values_list('email', flat=True))
                for row in ws.iter_rows(min_row=2, max_row=max_row):
                    email_cell = row[8]
                    result_cell = row[9]
                    email = str(email_cell.value).strip() if email_cell.value else ''
                    if email in yes_emails:
                        result_cell.value = 'Yes'
                    else:
                        result_cell.value = ''
                export_time = timezone.localtime()
                export_filename = f"tu_pot_membership_{export_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                export_path = f"media/tu_pot_exports/{export_filename}"
                wb.save(export_path)
                TUPOTExportHistory.objects.create(filename=export_filename, user=request.user if request.user.is_authenticated else None)
                request.session['pot_download_url'] = f"/media/tu_pot_exports/{export_filename}"
                request.session['pot_export_time'] = export_time.strftime('%Y-%m-%d %H:%M:%S')
                # 2. Newcomer file
                from openpyxl import Workbook
                out_wb = Workbook()
                out_ws = out_wb.active
                headers = [
                    'username', 'password', 'email', 'person_number', 'full_name_en', 'full_name_vn', 'dob', 'gender', 'discipline', 'job_title',
                    'floor', 'working_type', 'identity_number', 'native_place', 'ethnicity', 'religion', 'education_level', 'specialization',
                    'address', 'trade_union_member', 'membership_type_by_admin', 'membership_since', 'children'
                ]
                out_ws.append(headers)
                current_emails = set(Employee.objects.values_list('email', flat=True))
                for row in ws.iter_rows(min_row=2, max_row=max_row):
                    person_number = str(row[1].value).strip() if row[1].value else ''
                    full_name_vn = str(row[2].value).strip() if row[2].value else ''
                    date_joining = row[4].value
                    email = str(row[8].value).strip() if row[8].value else ''
                    if email and email not in current_emails:
                        username = email.split('@')[0] if '@' in email else ''
                        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                        if date_joining:
                            try:
                                month_str = date_joining.strftime('%b')
                            except Exception:
                                from datetime import datetime
                                try:
                                    date_obj = datetime.strptime(str(date_joining), '%Y-%m-%d')
                                    month_str = date_obj.strftime('%b')
                                except Exception:
                                    month_str = 'Month'
                            membership_type = f'Newcomer ({month_str})'
                            membership_since = str(date_joining)
                        else:
                            membership_type = 'Newcomer (Month)'
                            membership_since = ''
                        out_ws.append([
                            username, password, email, person_number, full_name_vn, full_name_vn, '1/1/1900', 'No Data', 'No Data', 'No Data',
                            0, 'No Data', person_number, 'No Data', 'No Data', 'No Data', 'No Data', 'No Data',
                            'No Data', 'No', membership_type, membership_since, '[]'
                        ])
                tu_export_time = timezone.localtime()
                tu_export_filename = f"tu_pot_newcomer_{tu_export_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                tu_export_path = f"media/tu_pot_exports/{tu_export_filename}"
                out_wb.save(tu_export_path)
                tu_download_url = f"/media/tu_pot_exports/{tu_export_filename}"
                TUPOTExportHistory.objects.create(filename=tu_export_filename, user=request.user if request.user.is_authenticated else None)
                request.session['tu_download_url'] = tu_download_url
                request.session['tu_export_time'] = tu_export_time.strftime('%Y-%m-%d %H:%M:%S')
                # 3. Missing file
                uploaded_emails = set()
                extra_info_map = {}
                for row in ws.iter_rows(min_row=2, max_row=max_row):
                    email = str(row[8].value).strip() if row[8].value else ''
                    extra_info = str(row[6].value).strip() if row[6].value else ''
                    if email:
                        uploaded_emails.add(email)
                        if extra_info:
                            extra_info_map[email] = extra_info
                all_employees = Employee.objects.values_list('email', flat=True)
                missing_wb = Workbook()
                missing_ws = missing_wb.active
                missing_ws.append(['Email', 'Extra Info'])
                for emp_email in all_employees:
                    if emp_email not in uploaded_emails:
                        info = extra_info_map.get(emp_email, '')
                        missing_ws.append([emp_email, info])
                missing_filename = f"tu_pot_missing_{tu_export_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                missing_path = f"media/tu_pot_exports/{missing_filename}"
                missing_wb.save(missing_path)
                request.session['tu_missing_download_url'] = f"/media/tu_pot_exports/{missing_filename}"
                TUPOTExportHistory.objects.create(filename=missing_filename, user=request.user if request.user.is_authenticated else None)
                # 4. Resign file
                resign_wb = Workbook()
                resign_ws = resign_wb.active
                resign_ws.append(['Email', 'Extra Info'])
                for row in ws.iter_rows(min_row=2, max_row=max_row):
                    email = str(row[8].value).strip() if row[8].value else ''
                    extra_info = str(row[6].value).strip() if row[6].value else ''
                    if email and extra_info:
                        resign_ws.append([email, extra_info])
                resign_filename = f"tu_pot_resign_{tu_export_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
                resign_path = f"media/tu_pot_exports/{resign_filename}"
                resign_wb.save(resign_path)
                request.session['tu_resign_download_url'] = f"/media/tu_pot_exports/{resign_filename}"
                TUPOTExportHistory.objects.create(filename=resign_filename, user=request.user if request.user.is_authenticated else None)
                return redirect('tu_pot')
        except Exception as e:
            error = f'Error processing file: {e}'
    
    is_superuser = request.user.is_superuser if request.user.is_authenticated else False
    is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
    is_pot = request.user.groups.filter(name='pot').exists() if request.user.is_authenticated else False
    from django.utils import timezone
    now = timezone.localtime()
    if export_filename:
        download_url = f"/media/tu_pot_exports/{export_filename}"
    # Get history for display
    history = TUPOTExportHistory.objects.order_by('-export_time')[:20]
    membership_history = TUPOTExportHistory.objects.filter(filename__startswith='tu_pot_membership').order_by('-export_time')
    # Get POT download info from session (after redirect)
    pot_download_url = request.session.pop('pot_download_url', None)
    pot_export_time = request.session.pop('pot_export_time', None)
    # Get TU download info from session (after redirect)
    tu_download_url = request.session.pop('tu_download_url', tu_download_url)
    tu_export_time = request.session.pop('tu_export_time', tu_export_time)
    tu_missing_download_url = request.session.pop('tu_missing_download_url', None)
    tu_resign_download_url = request.session.pop('tu_resign_download_url', None)
    return render(request, 'employee/tu_pot.html', {
        'error': error,
        'is_superuser': is_superuser,
        'is_committee': is_committee,
        'is_pot': is_pot,
        'now': now,
        'download_url': download_url,
        'export_time': export_time,
        'history': history,
        'membership_history': membership_history,
        'tu_download_url': tu_download_url,
        'tu_export_time': tu_export_time,
        'pot_download_url': pot_download_url,
        'pot_export_time': pot_export_time,
        'tu_missing_download_url': tu_missing_download_url,
        'tu_resign_download_url': tu_resign_download_url,
    })
