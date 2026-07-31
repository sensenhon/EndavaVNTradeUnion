

import io
import os
import json
import pandas as pd
import openpyxl
import datetime
from lunardate import LunarDate
from datetime import date
from openpyxl import load_workbook
from django import forms
from django.urls import reverse
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash, authenticate, login as auth_login
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.forms import modelformset_factory, inlineformset_factory, modelformset_factory
from django.db.models import Q, Sum
from django.views.decorators.http import require_POST, require_GET
from django.conf import settings
from .models import Employee, EditHistory, Discipline, Floor, EditHistory, Employee, Children, TUCommittee, EmployeeGiftYear, FinancialCategory, TUFinancialTransaction, FinancialDescription, ClubFinancialTransaction, Club, FinancialOpeningBalance
from .forms import EmployeeRegisterForm, EmployeeLoginForm, EmployeeRegisterForm, ClubFinancialForm

# Dùng chung cho dashboard và export
DISPLAY_FIELDS = [
	('person_number', 'Employee ID'),
	('user', 'Username'),
	('full_name_en', 'Full name (EN)'),
	('full_name_vn', 'Full name (VN)'),
	('email', 'Email'),
	('gender', 'Gender'),
	('dob', 'Birth Date'),
	('birth_month', 'Birth Month'),
	('discipline', 'Discipline'),
	('floor', 'Floor'),
	('job_title', 'Position'),
	('working_type', 'Work Type'),
	('identity_number', 'Identity Number'),
	('native_place', 'Native Place'),
	('ethnicity', 'Ethnicity'),
	('religion', 'Religion'),
	('education_level', 'Education Level'),
	('specialization', 'Specialization'),
	('address', 'Address'),
	('trade_union_member', 'Trade Union Member'),
]

def is_committee_or_superuser(user):
	return user.is_superuser or user.groups.filter(name='TU committee').exists()

def is_clubadmin(user):
	return user.groups.filter(name='clubadmin').exists()

def is_superuser_committee_clubadmin(user):
	return user.is_superuser or user.groups.filter(name='TU committee').exists() or is_clubadmin(user)

@login_required
@user_passes_test(is_superuser_committee_clubadmin)
def club_financial(request):
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	is_pot = request.user.groups.filter(name='pot').exists() if request.user.is_authenticated else False
	user_groups = request.user.groups.values_list('name', flat=True)
	clubadmin_group = None
	club_obj = None
	year = request.GET.get('year')
	month = request.GET.get('month')
	now = timezone.now()
	year = int(year) if year else now.year
	if month in [None, '', 'None']:
		month = None
	else:
		try:
			month = int(month)
		except Exception:
			month = None
	for g in user_groups:
		if g.lower().startswith('clubadmin'):
			clubadmin_group = g
			break
	if clubadmin_group:
		club_name = clubadmin_group.replace('ClubAdmin ', '').strip()
		club_obj = Club.objects.filter(name=club_name).first()
	club_list = list(Club.objects.all())
	if is_superuser:
		club_id = request.GET.get('club')
		if club_id in [None, '', 'None']:
			club_id = club_list[0].id if club_list else None
		club_obj = Club.objects.filter(id=club_id).first() if club_id else None
		show_club_dropdown = True
	else:
		show_club_dropdown = False
	# Xử lý submit transaction
	if request.method == 'POST':
		post_data = request.POST.copy()
		if not is_superuser and club_obj:
			post_data['club'] = club_obj.id
		form = ClubFinancialForm(post_data, request.FILES, initial={'request': request})
		if form.is_valid():
			tx = form.save(commit=False)
			tx.created_by = Employee.objects.filter(user=request.user).first()
			tx.save()
			messages.success(request, 'Transaction submitted successfully!')
			return redirect(request.path + f'?club={club_obj.id if club_obj else ''}&year={year}&month={month}')
	else:
		form = ClubFinancialForm(initial={'request': request})
	transactions = ClubFinancialTransaction.objects.filter(date__year=year)
	if month:
		transactions = transactions.filter(date__month=month)
	if club_obj:
		transactions = transactions.filter(club=club_obj)
	opening_obj = transactions.order_by('date').first()
	opening_balance = opening_obj.opening_balance if opening_obj and hasattr(opening_obj, 'opening_balance') else 0
	total_income = sum(t.amount for t in transactions if t.financial_type == 'income')
	total_expense = sum(t.amount for t in transactions if t.financial_type == 'expense')
	closing_balance = opening_balance + total_income - total_expense
	club_summary = {
		'year': year,
		'month': month,
		'club_name': club_obj.name if club_obj else '',
		'opening_balance': opening_balance,
		'total_income': total_income,
		'total_expense': total_expense,
		'closing_balance': closing_balance
	}
	# recent_transactions: luôn filter theo club_obj (superuser chọn club, clubadmin chỉ thấy club mình)
	recent_transactions = ClubFinancialTransaction.objects.filter(club=club_obj).order_by('-date', '-created_at')[:10] if club_obj else []
	category_list = [
		{
			'id': cat.id,
			'name': cat.name,
			'type': cat.type,
		} for cat in FinancialCategory.objects.filter(for_type__in=['club', 'both'])
	]
	description_list = [
		{
			'id': desc.id,
			'description': desc.description,
			'category_id': desc.category_id,
		} for desc in FinancialDescription.objects.all()
	]
	return render(request, 'employee/club_financial.html', {
		'form': form,
		'club_summary': club_summary,
		'club_list': club_list,
		'show_club_dropdown': show_club_dropdown,
		'year': year,
		'month': month,
		'is_superuser': is_superuser,
		'is_committee': is_committee,
		'is_pot': is_pot,
		'is_clubadmin': bool(clubadmin_group),
		'recent_transactions': recent_transactions,
		'category_list': category_list,
		'description_list': description_list
	})

def get_children_info(employee_qs, june_first):
	result = {}
	for emp in employee_qs:
		children_with_age = []
		for child in emp.children.all():
			if child.dob:
				age = june_first.year - child.dob.year - ((june_first.month, june_first.day) < (child.dob.month, child.dob.day))
			else:
				age = None
			children_with_age.append({'child': child, 'age': age})
		result[emp.id] = children_with_age
	return result

def get_children_autumn_gift_info(emp, autumn_date):
	result = {}
	for emp_item in emp:
		children_list = emp_item.children.all() if hasattr(emp_item, 'children') else []
		children_with_age = []
		for child in children_list:
			if child.dob:
				age = autumn_date.year - child.dob.year - ((autumn_date.month, autumn_date.day) < (child.dob.month, child.dob.day))
			else:
				age = None
			children_with_age.append({'child': child, 'age': age})
		result[emp_item.id] = children_with_age
	return result

def build_summary(year, type_name):
		categories = FinancialCategory.objects.filter(type=type_name).order_by('code')
		summary = []
		for cat in categories:
			descriptions = FinancialDescription.objects.filter(category=cat)
			desc_rows = []
			total_spent = 0
			estimated_cat = cat.estimated_expense if cat.estimated_expense else 0
			for desc in descriptions:
				amount = TUFinancialTransaction.objects.filter(date__year=year, category=cat, description=desc).aggregate(total=Sum('amount'))['total']
				if not amount:
					amount = 0
				estimated_expense_description = desc.estimated_expense if hasattr(desc, 'estimated_expense') and desc.estimated_expense is not None else 0
				desc_percentage = (amount / estimated_expense_description * 100) if estimated_expense_description and estimated_expense_description > 0 else 0
				desc_rows.append({
					'desc': str(desc),
					'amount': amount,
					'desc_percentage': round(desc_percentage, 2),
					'estimated_expense_description': estimated_expense_description
				})
				total_spent += amount
			percent = (total_spent / estimated_cat * 100) if estimated_cat and estimated_cat > 0 else 0
			summary.append({
				'category': str(cat),
				'descriptions': desc_rows,
				'estimated_expense_category': estimated_cat,
				'total_amount': total_spent,
				'percentage': round(percent, 2)
			})
		return summary

@user_passes_test(is_committee_or_superuser)
@user_passes_test(is_committee_or_superuser)
def committee_dashboard(request):
	display_fields = DISPLAY_FIELDS
	if request.user.groups.filter(name='TU committee').exists():
		hidden_fields = [
			'identity_number', 'native_place', 'job_title', 'ethnicity', 'religion', 'education_level', 'specialization', 'address', 'dob'
		]
		display_fields = [f for f in DISPLAY_FIELDS if f[0] not in hidden_fields]

	# Query all employees data
	employees = Employee.objects.all()
	newcomers = Employee.objects.filter(membership_type_by_admin__name__icontains="newcomer")
	name_query_newcomer = request.GET.get('name', '').strip()
	if name_query_newcomer:
		newcomers = newcomers.filter(
			Q(full_name_en__icontains=name_query_newcomer) |
			Q(user__username__icontains=name_query_newcomer)
		)

	# Query all employees with "withdrawn" or "no" and create filter
	withdrawn_no = Employee.objects.filter(
		Q(membership_type_by_admin__name__icontains="withdrawn") |
		Q(membership_type_by_admin__name__icontains="no")
	)
	name_query_withdrawn_no = request.GET.get('name', '').strip()
	if name_query_withdrawn_no:
		withdrawn_no = withdrawn_no.filter(
			Q(full_name_en__icontains=name_query_withdrawn_no) |
			Q(user__username__icontains=name_query_withdrawn_no)
		)

	# Query all employees with "resignation" and create filter
	resignation = Employee.objects.filter(membership_type_by_admin__name__icontains="resignation")
	name_query_resignation = request.GET.get('name', '').strip()
	if name_query_resignation:
		resignation = resignation.filter(
			Q(full_name_en__icontains=name_query_resignation) |
			Q(user__username__icontains=name_query_resignation)
		)
	
	# Query all employees with "maternity" and create filter
	maternity = Employee.objects.filter(membership_type_by_admin__name__icontains="maternity")
	name_query_maternity = request.GET.get('name', '').strip()
	if name_query_maternity:
		maternity = maternity.filter(
			Q(full_name_en__icontains=name_query_maternity) |
			Q(user__username__icontains=name_query_maternity)
		)

	# Query all employees with "military" and create filter
	military = Employee.objects.filter(membership_type_by_admin__name__icontains="military")
	name_query_military = request.GET.get('name', '').strip()
	if name_query_military:
		military = military.filter(
			Q(full_name_en__icontains=name_query_military) |
			Q(user__username__icontains=name_query_military)
		)

	# Exclude the employees with above membership types
	employees = employees.exclude(
		Q(membership_type_by_admin__name__icontains="newcomer") |
		Q(membership_type_by_admin__name__icontains="withdrawn") |
		Q(membership_type_by_admin__name__icontains="resignation") |
		Q(membership_type_by_admin__name__icontains="maternity") |
		Q(membership_type_by_admin__name__icontains="military")
	)

	# Employee list with "YES"
	employees = employees.filter(membership_type_by_admin__name__iexact="Yes")
	
	tu_committees = TUCommittee.objects.all()
	tu_committee_query = request.GET.get('tu_committee', '').strip()
	if tu_committee_query:
		# Lọc employee theo TUCommittee phụ trách
		filtered_ids = []
		for emp in employees:
			committee = None
			if str(emp.floor) == '0':
				committee = TUCommittee.objects.filter(position='President').first()
			else:
				committee = TUCommittee.objects.filter(responsible_floor=str(emp.floor)).first()
			if not committee:
				committee = TUCommittee.objects.filter(position='Vice President').first()
			if committee and str(committee.id) == tu_committee_query:
				filtered_ids.append(emp.id)
		employees = employees.filter(id__in=filtered_ids)
	discipline_list = Discipline.objects.all().order_by('name')
	floor_list = Floor.objects.all().order_by('name')

	# Filtering
	name_query = request.GET.get('name', '').strip()
	discipline_query = request.GET.get('discipline', '').strip()
	floor_query = request.GET.get('floor', '').strip()
	birth_month_query = request.GET.getlist('birth_month')
	sort_field = request.GET.get('sort', '')
	birth_month_query = [m for m in birth_month_query if m]
	if name_query:
		employees = employees.filter(
			full_name_en__icontains=name_query
		) | employees.filter(
			user__username__icontains=name_query
		)
	if discipline_query:
		employees = employees.filter(discipline__name__icontains=discipline_query)
	if floor_query:
		employees = employees.filter(floor__name__iexact=floor_query)
	if birth_month_query: 
		employees = employees.filter(dob__month__in=birth_month_query) 

	# Sorting
	valid_sort_fields = [f[0] for f in display_fields if f[0] != 'birth_month']
	if sort_field and sort_field in valid_sort_fields:
		employees = employees.order_by(sort_field)
	else:
		employees = employees.order_by('dob__month')
	histories = EditHistory.objects.all().order_by('-edit_time')
	is_committee = request.user.groups.filter(name='TU committee').exists()

	# Map employee id to TUCommittee (user, email) by floor for main employees
	tu_committee_map = {}
	for emp in employees:
		committee = None
		if emp.floor and emp.floor.name == '0':
			committee = TUCommittee.objects.filter(position='President').first()
		else:
			committee = TUCommittee.objects.filter(responsible_floor=emp.floor.name if emp.floor else '').first()
		if committee:
			tu_committee_map[emp.id] = f"{committee.user.username} ({committee.email})"
		else:
			# Fallback: Vice President
			vp = TUCommittee.objects.filter(position='Vice President').first()
			if vp:
				tu_committee_map[emp.id] = f"{vp.user.username} ({vp.email})"
			else:
				tu_committee_map[emp.id] = "-"

	# Map TUCommittee for Newcomers: always assign President
	tu_committee_map_newcomers = {}
	president = TUCommittee.objects.filter(position='President').first()
	for emp in newcomers:
		if president:
			tu_committee_map_newcomers[emp.id] = f"{president.user.username} ({president.email})"
		else:
			tu_committee_map_newcomers[emp.id] = "-"
	

	# Prepare children age data for June gift eligibility
	current_year = date.today().year
	june_first = date(current_year, 6, 1)
	employee_children_ages = get_children_info(employees, june_first)
	newcomer_children_ages = get_children_info(newcomers, june_first)
	withdrawn_no_children_ages = get_children_info(withdrawn_no, june_first)
	resignation_children_ages = get_children_info(resignation, june_first)
	maternity_children_ages = get_children_info(maternity, june_first)
	military_children_ages = get_children_info(military, june_first)

	# Helper: count ticked June gift boxes per employee
	def count_june_gift_checked(children_ages_dict):
		result = {}
		for emp_id, children_list in children_ages_dict.items():
			result[emp_id] = sum(1 for info in children_list if getattr(info['child'], 'june_gift_received', False))
		return result

	employee_june_gift_checked = count_june_gift_checked(employee_children_ages)
	newcomer_june_gift_checked = count_june_gift_checked(newcomer_children_ages)
	withdrawn_no_june_gift_checked = count_june_gift_checked(withdrawn_no_children_ages)
	resignation_june_gift_checked = count_june_gift_checked(resignation_children_ages)
	maternity_june_gift_checked = count_june_gift_checked(maternity_children_ages)
	military_june_gift_checked = count_june_gift_checked(military_children_ages)

	# Prepare children age data for Autumn gift eligibility
	current_year = date.today().year
	autumn_date = LunarDate(current_year, 8, 15).toSolarDate()
	employee_children_autumn_ages = get_children_autumn_gift_info(employees, autumn_date)
	newcomer_children_autumn_ages = get_children_autumn_gift_info(newcomers, autumn_date)
	withdrawn_no_children_autumn_ages = get_children_autumn_gift_info(withdrawn_no, autumn_date)
	resignation_children_autumn_ages = get_children_autumn_gift_info(resignation, autumn_date)
	maternity_children_autumn_ages = get_children_autumn_gift_info(maternity, autumn_date)
	military_children_autumn_ages = get_children_autumn_gift_info(military, autumn_date)

	return render(request, 'employee/committee_dashboard.html', {
		'employees': employees,
		'newcomers': newcomers,
		'withdrawn_no': withdrawn_no,
		'resignation': resignation,
		'maternity': maternity,
		'military': military,
		'histories': histories,
		'display_fields': display_fields,
		'discipline_list': discipline_list,
		'floor_list': floor_list,
		'birth_month_options': range(1, 13),
		'selected_birth_month': birth_month_query,
		'is_superuser': request.user.is_superuser,
		'is_committee': is_committee,
		'tu_committee_map': tu_committee_map,
		'tu_committee_map_newcomers': tu_committee_map_newcomers,
		'tu_committees': tu_committees,
		'selected_tu_committee': tu_committee_query,
		'is_pot': request.user.groups.filter(name='pot').exists() if request.user.is_authenticated else False,
		'employee_children_ages': employee_children_ages,
		'newcomer_children_ages': newcomer_children_ages,
		'withdrawn_no_children_ages': withdrawn_no_children_ages,
		'resignation_children_ages': resignation_children_ages,
		'maternity_children_ages': maternity_children_ages,
		'military_children_ages': military_children_ages,
		'employee_children_autumn_ages': employee_children_autumn_ages,
		'newcomer_children_autumn_ages': newcomer_children_autumn_ages,
		'withdrawn_no_children_autumn_ages': withdrawn_no_children_autumn_ages,
		'resignation_children_autumn_ages': resignation_children_autumn_ages,
		'maternity_children_autumn_ages': maternity_children_autumn_ages,
		'military_children_autumn_ages': military_children_autumn_ages,
		'employee_june_gift_checked': employee_june_gift_checked,
		'newcomer_june_gift_checked': newcomer_june_gift_checked,
		'withdrawn_no_june_gift_checked': withdrawn_no_june_gift_checked,
		'resignation_june_gift_checked': resignation_june_gift_checked,
		'maternity_june_gift_checked': maternity_june_gift_checked,
		'military_june_gift_checked': military_june_gift_checked,
	})

# Export dashboard to Excel (filtered)
@user_passes_test(is_committee_or_superuser)
def export_dashboard_excel(request):
	employees = Employee.objects.all()
	name_query = request.GET.get('name', '').strip()
	discipline_query = request.GET.get('discipline', '').strip()
	floor_query = request.GET.get('floor', '').strip()
	birth_month_query = [m for m in request.GET.getlist('birth_month') if m]
	sort_field = request.GET.get('sort', '')
	if name_query:
		employees = employees.filter(full_name_en__icontains=name_query)
	if discipline_query:
		employees = employees.filter(discipline__name__icontains=discipline_query)
	if floor_query:
		employees = employees.filter(floor__name__iexact=floor_query)
	if birth_month_query:
		employees = employees.filter(dob__month__in=birth_month_query)
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	# only for superuser
	hidden_fields = [
		'dob', 'identity_number', 'native_place', 'ethnicity', 'religion',
		'education_level', 'specialization', 'address'
	]
	display_fields = [f for f in DISPLAY_FIELDS if is_superuser or f[0] not in hidden_fields]
	extra_fields = [
		('membership_type_by_admin', 'MembershipByTU'),
		('membership_since', 'MembershipSince'),
		('tu_committee', 'TUCommittee'),
		('birthday_gift_received', 'Birthday Gift'),
		('mooncake_gift_received', 'Mooncake Gift'),
		('tet_gift_received', 'Tet Gift'),
		('luckymoney_gift_received', 'Lucky Money Gift'),
		('children_gift', 'Children Gift'),
		('children_gift_autumn', 'Children Gift (Autumn)'),
		('children', 'Children'),
	]
	
	current_year = datetime.date.today().year
	autumn_date = LunarDate(current_year, 8, 15).toSolarDate()

	valid_sort_fields = [f[0] for f in display_fields]
	if sort_field and sort_field in valid_sort_fields:
		employees = employees.order_by(sort_field)
	else:
		employees = employees.order_by('dob__month')
	# Prepare data for Excel
	data = []
	for emp in employees:
		row = []
		for field, label in display_fields:
			if field == 'user':
				row.append(emp.user.username if emp.user else '')
			elif field == 'gender':
				row.append(emp.gender.name if emp.gender else '')
			elif field == 'discipline':
				row.append(emp.discipline.name if emp.discipline else '')
			elif field == 'floor':
				row.append(emp.floor.name if emp.floor else '')
			elif field == 'job_title':
				row.append(emp.job_title.name if emp.job_title else '')
			elif field == 'working_type':
				row.append(emp.working_type.name if emp.working_type else '')
			elif field == 'dob':
				row.append(emp.dob.strftime('%Y-%m-%d') if emp.dob else '')
			elif field == 'birth_month':
				row.append(str(emp.dob.month) if emp.dob else '')
			elif field == 'tu_committee':
				committee = getattr(emp, 'tu_committee', None)
				if committee:
					row.append(str(committee))
				else:
					row.append('')
			else:
				row.append(getattr(emp, field, ''))
		# MembershipByTU
		row.append(emp.membership_type_by_admin.name if hasattr(emp, 'membership_type_by_admin') and emp.membership_type_by_admin else '')
		# MembershipSince
		row.append(emp.membership_since.strftime('%Y-%m-%d') if hasattr(emp, 'membership_since') and emp.membership_since else '')
		# TUCommittee
		committee = getattr(emp, 'tu_committee', None)
		if committee:
			row.append(str(committee))
		else:
			row.append('')
		# Birthday Gift
		row.append('Yes' if emp.birthday_gift_received else 'No')
		# Mooncake Gift
		row.append('Yes' if emp.mooncake_gift_received else 'No')
		# Tet Gift
		row.append('Yes' if emp.tet_gift_received else 'No')
		# Lucky Money Gift
		row.append('Yes' if emp.luckymoney_gift_received else 'No')
		# Children Gift (June)
		children_list = emp.children.all() if hasattr(emp, 'children') else []
		if children_list:
			children_gift_str = '; '.join([f"{child.name}: {'Yes' if child.june_gift_received else 'No'}" for child in children_list])
			row.append(children_gift_str)
		else:
			row.append('')
		# Children Gift (Autumn)
		if children_list:
			autumn_gift_str = [f"{child.name}: {'Yes' if child.autumn_gift_received else 'No'}" for child in children_list]
			row.append('; '.join(autumn_gift_str))
		else:
			row.append('')
		# Children
		if children_list:
			children_str = '; '.join([f"{child.name} ({child.dob})" for child in children_list])
			row.append(children_str)
		else:
			row.append('')
		data.append(row)
	df = pd.DataFrame(data, columns=[label for field, label in display_fields] + [label for field, label in extra_fields])
	output = io.BytesIO()
	with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
		df.to_excel(writer, index=False, sheet_name='Dashboard')
	output.seek(0)
	response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=dashboard.xlsx'
	return response

def logout_view(request):
	logout(request)
	return redirect('home')

@require_GET
def statistics_view(request):
	# Get year from GET param, default to current year
	year = int(request.GET.get('year', datetime.date.today().year))

	# Total received gifts for employees
	total_birthday = EmployeeGiftYear.objects.filter(year=year, gift_type='birthday', received=True).count()
	total_mooncake = EmployeeGiftYear.objects.filter(year=year, gift_type='mooncake', received=True).count()
	total_tet = EmployeeGiftYear.objects.filter(year=year, gift_type='tet', received=True).count()
	total_luckymoney = EmployeeGiftYear.objects.filter(year=year, gift_type='luckymoney', received=True).count()

	# Children gifts: count by child
	total_june_children = Children.objects.filter(june_gift_received=True, employee__employeegiftyear__year=year).count()
	total_autumn_children = Children.objects.filter(autumn_gift_received=True, employee__employeegiftyear__year=year).count()

	# For year selection dropdown
	all_years = EmployeeGiftYear.objects.values_list('year', flat=True).distinct().order_by('-year')

	# Total permanent members for selected year (membership_type_by_admin.name == 'Yes' and membership_since <= year)
	total_permanent = Employee.objects.filter(
		membership_type_by_admin__name__iexact='Yes',
		membership_since__year__lte=year
	).count()

	# User info for navbar
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False

	return render(request, 'employee/statistics.html', {
		'selected_year': year,
		'all_years': all_years,
		'total_birthday': total_birthday,
		'total_mooncake': total_mooncake,
		'total_tet': total_tet,
		'total_luckymoney': total_luckymoney,
		'total_june_children': total_june_children,
		'total_autumn_children': total_autumn_children,
		'total_permanent': total_permanent,
		'is_superuser': is_superuser,
		'is_committee': is_committee,
	})

@login_required
def edit_children(request):
	emp_id = request.GET.get('id') or request.POST.get('id')
	# Luôn ưu tiên lấy employee theo id nếu có, cho mọi trường hợp
	if emp_id:
		employee = Employee.objects.get(id=emp_id)
	else:
		employee = Employee.objects.get(user=request.user)

	ChildrenFormSet = inlineformset_factory(
		Employee,
		Children,
		fields=('name', 'dob'),
		extra=1,
		can_delete=True,
		widgets={
			'dob': forms.DateInput(attrs={'type': 'date', 'class': 'form-control', 'placeholder': 'YYYY-MM-DD'})
		}
	)

	if request.method == 'POST':
		# Đảm bảo luôn lấy đúng employee theo id trong POST
		emp_id_post = request.POST.get('id')
		if emp_id_post:
			employee = Employee.objects.get(id=emp_id_post)
		formset = ChildrenFormSet(request.POST, instance=employee)
		if formset.is_valid():
			old_children = list(employee.children.all())
			old_map = {c.id: c for c in old_children}
			children = formset.save(commit=False)
			changes = []
			# Detect add/update
			for child in children:
				child.employee = employee
				if child.id is None:
					changes.append(f"Added child: {child.name} ({child.dob})")
				else:
					old = old_map.get(child.id)
					if old and (old.name != child.name or old.dob != child.dob):
						changes.append(f"Updated child: {old.name} ({old.dob}) -> {child.name} ({child.dob})")
				child.save()
			# Detect delete
			for obj in formset.deleted_objects:
				changes.append(f"Deleted child: {obj.name} ({obj.dob})")
				obj.delete()
			formset.save()
			# Ghi lịch sử nếu có thay đổi
			if changes:
				EditHistory.objects.create(
					employee=employee,
					edited_by=request.user,
					changes='; '.join(changes)
				)
			print("Formset valid, redirecting to profile")
			return redirect(f'/profile/?id={employee.id}')
		else:
			print("Formset errors:", formset.errors)
	else:
		formset = ChildrenFormSet(instance=employee)
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	return render(request, 'employee/edit_children.html', {'formset': formset, 'is_superuser': is_superuser, 'is_committee': is_committee, 'employee': employee})

@login_required
def change_password(request):
	if request.method == 'POST':
		form = PasswordChangeForm(request.user, request.POST)
		if form.is_valid():
			user = form.save()
			update_session_auth_hash(request, user)  # Giữ đăng nhập
			return redirect('profile')
	else:
		form = PasswordChangeForm(request.user)
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	return render(request, 'employee/change_password.html', {'form': form, 'is_superuser': is_superuser, 'is_committee': is_committee})

@login_required
def edit_profile(request):
	# Ẩn thông tin nhạy cảm nếu là TU committee chỉnh employee khác
	# Hide by both field names and display labels
	sensitive_fields = ['identity_number', 'native_place', 'ethnicity', 'religion', 'education_level', 'specialization', 'address']
	field_labels = {
		'identity_number': 'Identity Number',
		'native_place': 'Native Place',
		'ethnicity': 'Ethnicity',
		'religion': 'Religion',
		'education_level': 'Education Level',
		'specialization': 'Specialization',
		'address': 'Address',
	}
	hidden_fields = sensitive_fields + [field_labels[f] for f in sensitive_fields]
	show_limited = False
	emp_id = request.GET.get('id') or request.POST.get('id')
	if request.user.groups.filter(name='TU committee').exists() and emp_id:
		try:
			user_employee_id = Employee.objects.get(user=request.user).id
			if int(emp_id) != user_employee_id:
				show_limited = True
		except Employee.DoesNotExist:
			show_limited = True
	if (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()) and emp_id:
		employee = Employee.objects.get(id=emp_id)
	else:
		employee = Employee.objects.get(user=request.user)
	class EmployeeUpdateForm(EmployeeRegisterForm):
		class Meta(EmployeeRegisterForm.Meta):
			exclude = ['username', 'password']
		def __init__(self, *args, **kwargs):
			self.show_limited = kwargs.pop('show_limited', False)
			super().__init__(*args, **kwargs)
			for field in ['username', 'password']:
				if field in self.fields:
					self.fields.pop(field)
			# Nếu là employee tự chỉnh thì disable 3 field và membership_since
			if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
				for field in ['full_name_en', 'full_name_vn', 'email', 'membership_since']:
					if field in self.fields:
						self.fields[field].disabled = True
			# Nếu show_limited thì các trường nhạy cảm không required
			if show_limited:
				for field in sensitive_fields:
					if field in self.fields:
						self.fields[field].required = False
				# TU committee editing another profile: replace dob with month/day inputs
				if 'dob' in self.fields:
					self.fields.pop('dob')
					self.fields['dob_month'] = forms.IntegerField(
						required=True,
						min_value=1,
						max_value=12,
						label='Birth Month',
						widget=forms.NumberInput(attrs={'class': 'form-control'})
					)
					self.fields['dob_day'] = forms.IntegerField(
						required=True,
						min_value=1,
						max_value=31,
						label='Birth Day',
						widget=forms.NumberInput(attrs={'class': 'form-control'})
					)
					if self.instance and getattr(self.instance, 'dob', None):
						self.fields['dob_month'].initial = self.instance.dob.month
						self.fields['dob_day'].initial = self.instance.dob.day
			# Only show membership_type_by_admin for superuser or TU committee
			if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
				if 'membership_type_by_admin' in self.fields:
					self.fields.pop('membership_type_by_admin')
		def clean(self):
			cleaned_data = super().clean()
			if self.show_limited and 'dob_month' in self.fields and 'dob_day' in self.fields:
				month = cleaned_data.get('dob_month')
				day = cleaned_data.get('dob_day')
				if month is not None and day is not None:
					try:
						date(2000, month, day)
					except ValueError:
						raise forms.ValidationError('Invalid Birth Month and Day combination.')
			return cleaned_data
	if request.method == 'POST':
		form = EmployeeUpdateForm(request.POST, instance=employee, show_limited=show_limited)
		print('Form errors:', form.errors)
		if form.is_valid():
			# Lưu dữ liệu cũ
			old_employee = Employee.objects.get(pk=employee.pk)
			# Nếu show_limited, preserve sensitive fields
			if show_limited:
				for field in sensitive_fields:
					setattr(form.instance, field, getattr(old_employee, field))
				if 'dob_month' in form.cleaned_data and 'dob_day' in form.cleaned_data:
					month = form.cleaned_data.get('dob_month')
					day = form.cleaned_data.get('dob_day')
					if month and day:
						old_year = old_employee.dob.year if old_employee.dob else date.today().year
						form.instance.dob = date(old_year, month, day)
			updated_employee = form.save()
			changes = []
			field_labels = {
				'full_name_en': 'Full name (EN)',
				'full_name_vn': 'Full name (VI)',
				'dob': 'Date of Birth',
				'gender': 'Gender',
				'discipline': 'Discipline',
				'job_title': 'Job Title',
				'floor': 'Floor',
				'working_type': 'Working Type',
				'identity_number': 'Identity Number',
				'native_place': 'Native Place',
				'ethnicity': 'Ethnicity',
				'religion': 'Religion',
				'education_level': 'Education Level',
				'specialization': 'Specialization',
				'address': 'Address',
				'trade_union_member': 'Trade Union member',
			}
			compare_fields = [
				'full_name_en', 'full_name_vn', 'dob', 'gender', 'discipline', 'job_title', 'floor',
				'working_type', 'identity_number', 'native_place', 'ethnicity', 'religion',
				'education_level', 'specialization', 'address', 'trade_union_member', 'membership_since'
			]
			for field in compare_fields:
				old_val = getattr(old_employee, field)
				new_val = getattr(updated_employee, field)
				# Nếu là FK thì lấy tên
				if hasattr(new_val, 'name'):
					new_val = new_val.name
				if hasattr(old_val, 'name'):
					old_val = old_val.name
				# Boolean cho trade_union_member
				if field == 'trade_union_member':
					old_val = 'Yes' if old_val else 'No'
					new_val = 'Yes' if new_val else 'No'
				# Format Membership Since
				if field == 'membership_since':
					def format_dt(dt):
						if isinstance(dt, datetime.datetime):
							return dt.strftime('%B, %Y')
						return str(dt) if dt else ''
					old_val = format_dt(old_val)
					new_val = format_dt(new_val)
				if old_val != new_val:
					label = field_labels.get(field, field)
					if field == 'membership_since':
						label = 'Membership Since'
					changes.append(f"{label}: '{old_val}' → '{new_val}'")
			if changes:
				changes_str = '; '.join(changes)
				print(f"EditHistory: {changes_str}")
				EditHistory.objects.create(
					employee=employee,
					edited_by=request.user,
					changes=changes_str
				)
			# Nếu là superuser hoặc TU committee thì redirect về /profile/?id=<employee_id>, còn lại thì về /profile/
			if request.user.is_superuser or request.user.groups.filter(name='TU committee').exists():
				return redirect(f'/profile/?id={employee.id}')
			else:
				return redirect('/profile/')
	else:
		form = EmployeeUpdateForm(instance=employee, show_limited=show_limited)
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	return render(request, 'employee/edit_profile.html', {
		'form': form,
		'is_superuser': is_superuser,
		'is_committee': is_committee,
		'show_limited': show_limited,
		'hidden_fields': hidden_fields,
		'employee': employee
	})

@login_required
def profile(request):
	emp_id = request.GET.get('id')
	if (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()) and emp_id:
		try:
			employee = Employee.objects.get(id=emp_id)
		except Employee.DoesNotExist:
			return render(request, 'employee/profile.html', {'error': 'Employee not found.'})
	else:
		try:
			employee = Employee.objects.get(user=request.user)
		except Employee.DoesNotExist:
			# Tự động tạo Employee cho superuser nếu chưa có
			if request.user.is_superuser:
				employee = Employee.objects.create(
					user=request.user,
					person_number=get_random_string(8),
					full_name_en=request.user.username,
					full_name_vn=request.user.username,
					email=request.user.email or f"{request.user.username}@example.com",
					dob="2000-01-01"
				)
			else:
				return render(request, 'employee/profile.html', {'error': 'Bạn chưa có hồ sơ nhân viên.'})
	ChildrenFormSet = modelformset_factory(Children, fields=('name', 'dob'), extra=1, can_delete=True)
	if request.method == 'POST':
		formset = ChildrenFormSet(request.POST, queryset=employee.children.all())
		if formset.is_valid():
			children = formset.save(commit=False)
			for child in children:
				child.employee = employee
				child.save()
			for obj in formset.deleted_objects:
				obj.delete()
		# Sau khi xử lý POST, reload lại dữ liệu
		formset = ChildrenFormSet(queryset=employee.children.all())
	else:
		formset = ChildrenFormSet(queryset=employee.children.all())
	children = employee.children.all()
	history = employee.edithistory_set.order_by('-edit_time')
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	# Ẩn thông tin nhạy cảm nếu là TU committee xem profile người khác
	sensitive_fields = ['identity_number', 'native_place', 'ethnicity', 'religion', 'education_level', 'specialization', 'address']
	field_labels = {
		'identity_number': 'Identity Number',
		'native_place': 'Native Place',
		'ethnicity': 'Ethnicity',
		'religion': 'Religion',
		'education_level': 'Education Level',
		'specialization': 'Specialization',
		'address': 'Address',
	}
	hidden_fields = sensitive_fields + [field_labels[f] for f in sensitive_fields]
	show_limited = False
	if request.user.groups.filter(name='TU committee').exists() and emp_id and int(emp_id) != request.user.employee.id:
		show_limited = True
	# Tìm TU Committee theo floor
	
	committee = None
	if employee.floor and str(employee.floor) == '0':
		committee = TUCommittee.objects.filter(position='President').first()
	elif employee.floor:
		committee = TUCommittee.objects.filter(responsible_floor=str(employee.floor)).first()
	if not committee:
		committee = TUCommittee.objects.filter(position='Vice President').first()
	tu_committee_display = f"{committee.user.username} - {committee.email}" if committee else "-"
	context = {
		'employee': employee,
		'formset': formset,
		'children': children,
		'history': history,
		'is_superuser': is_superuser,
		'is_committee': is_committee,
		'show_limited': show_limited,
		'hidden_fields': hidden_fields,
		'tu_committee_display': tu_committee_display,
	}
	if is_superuser or is_committee:
		context['membership_type_by_admin'] = employee.membership_type_by_admin if hasattr(employee, 'membership_type_by_admin') else None
	return render(request, 'employee/profile.html', context)

def login_view(request):
	if request.method == 'POST':
		form = EmployeeLoginForm(request.POST)
		if form.is_valid():
			username = form.cleaned_data['username']
			password = form.cleaned_data['password']
			user = authenticate(request, username=username, password=password)
			if user is not None:
				auth_login(request, user)
				return redirect('/home/')
			else:
				form.add_error(None, 'Username hoặc password không đúng.')
	else:
		form = EmployeeLoginForm()
	return render(request, 'employee/login.html', {'form': form})

def check_username(request):
	username = request.GET.get('username', '')
	exists = User.objects.filter(username=username).exists()
	return JsonResponse({'exists': exists})

def home(request):
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	is_pot = request.user.groups.filter(name='pot').exists() if request.user.is_authenticated else False
	return render(request, 'employee/home.html', {'is_superuser': is_superuser, 'is_committee': is_committee, 'is_pot': is_pot})

def register(request):
	class EmployeeRegisterFormNoMembership(EmployeeRegisterForm):
		def __init__(self, *args, **kwargs):
			super().__init__(*args, **kwargs)
			self.fields.pop('membership_type_by_admin', None)
			self.fields.pop('membership_since', None)

	if request.method == 'POST':
		form = EmployeeRegisterFormNoMembership(request.POST)
		if form.is_valid():
			form.save()
			messages.success(request, 'Đăng ký thành công!')
			return redirect('/home/')  # chuyển hướng tới trang Home
	else:
		form = EmployeeRegisterFormNoMembership()
	return render(request, 'employee/register.html', {'form': form})

@login_required
@require_POST
def update_birthday_gift(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    try:
        data = json.loads(request.body)
        emp_id = data.get('id')
        value = data.get('value')
        if value in [True, 'true', 'True', 1, '1']:
            checked = True
        else:
            checked = False
        emp = Employee.objects.get(id=emp_id)
        emp.birthday_gift_received = checked
        emp.save()
        # sync with EmployeeGiftYear
        year = date.today().year
        EmployeeGiftYear.objects.update_or_create(
            employee=emp,
            year=year,
            gift_type='birthday',
            defaults={'received': checked}
        )
        return JsonResponse({'success': True})
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def update_tet_gift(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    try:
        data = json.loads(request.body)
        emp_id = data.get('id')
        value = data.get('value')
        if value in [True, 'true', 'True', 1, '1']:
            checked = True
        else:
            checked = False
        emp = Employee.objects.get(id=emp_id)
        emp.tet_gift_received = checked
        emp.save()
        # sync with EmployeeGiftYear
        year = date.today().year
        EmployeeGiftYear.objects.update_or_create(
            employee=emp,
            year=year,
            gift_type='tet',
            defaults={'received': checked}
        )
        return JsonResponse({'success': True})
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
	
@login_required
@require_POST
def update_mooncake_gift(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    try:
        data = json.loads(request.body)
        emp_id = data.get('id')
        value = data.get('value')
        # Accept both boolean and string representations
        if value in [True, 'true', 'True', 1, '1']:
            checked = True
        else:
            checked = False
        emp = Employee.objects.get(id=emp_id)
        emp.mooncake_gift_received = checked
        emp.save()
        # sync with EmployeeGiftYear
        year = date.today().year
        EmployeeGiftYear.objects.update_or_create(
            employee=emp,
            year=year,
            gift_type='mooncake',
            defaults={'received': checked}
        )
        return JsonResponse({'success': True})
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def update_luckymoney_gift(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    try:
        data = json.loads(request.body)
        emp_id = data.get('id')
        value = data.get('value')
        # Accept both boolean and string representations
        if value in [True, 'true', 'True', 1, '1']:
            checked = True
        else:
            checked = False
        emp = Employee.objects.get(id=emp_id)
        emp.luckymoney_gift_received = checked
        emp.save()
        # sync with EmployeeGiftYear
        year = date.today().year
        EmployeeGiftYear.objects.update_or_create(
            employee=emp,
            year=year,
            gift_type='luckymoney',
            defaults={'received': checked}
        )
        return JsonResponse({'success': True})
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def update_june_gift(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    data = json.loads(request.body)
    child_id = data.get('child_id')
    received = data.get('june_gift_received')
    try:
        child = Children.objects.get(id=child_id)
        child.june_gift_received = received
        child.save()
        # sync with EmployeeGiftYear
        year = date.today().year
        EmployeeGiftYear.objects.update_or_create(
            employee=child.employee,
            year=year,
            gift_type='june',
            defaults={'received': received}
        )
        return JsonResponse({'success': True})
    except Children.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Child not found'})
	
@login_required
@require_POST
def update_autumn_gift(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    data = json.loads(request.body)
    child_id = data.get('child_id')
    received = data.get('autumn_gift_received')
    try:
        child = Children.objects.get(id=child_id)
        child.autumn_gift_received = received
        child.save()
        # sync with EmployeeGiftYear
        year = date.today().year
        EmployeeGiftYear.objects.update_or_create(
            employee=child.employee,
            year=year,
            gift_type='autumn',
            defaults={'received': received}
        )
        return JsonResponse({'success': True})
    except Children.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Child not found'})
	
class TUFinancialTransactionForm(forms.ModelForm):
	class Meta:
		model = TUFinancialTransaction
		fields = ['financial_type', 'category', 'date', 'payment_id', 'description', 'details', 'amount', 'payment_evidence']
		widgets = {
			'date': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
			'amount': forms.NumberInput(attrs={'class': 'form-control', 'min': 0}),
			'details': forms.Textarea(attrs={'class': 'form-control form-control-sm', 'rows': 2}),
		}

	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)
		for field in self.fields:
			self.fields[field].required = True
		self.fields['details'].widget.attrs['placeholder'] = 'Eg: F5 - sen.hon'
		self.fields['details'].widget.attrs['class'] = 'form-control form-control-sm'
		ftype = self.data.get('financial_type') or getattr(self.instance, 'financial_type', None)
		if ftype:
			self.fields['category'].queryset = FinancialCategory.objects.filter(type=ftype)
			self.fields['description'].queryset = FinancialDescription.objects.filter(type=ftype)
		else:
			self.fields['category'].queryset = FinancialCategory.objects.none()
			self.fields['description'].queryset = FinancialDescription.objects.none()

@login_required
def financial_view(request):
    is_superuser = request.user.is_superuser if request.user.is_authenticated else False
    is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
    if not (is_superuser or is_committee):
        return redirect('home')
    # Filter by year/month
    year = int(request.GET.get('year', timezone.now().year))
    month = request.GET.get('month')
    tu_summary = get_tu_financial_summary(year, month)
    transactions = TUFinancialTransaction.objects.filter(date__year=year)
    if month:
        transactions = transactions.filter(date__month=month)
    transactions = transactions.order_by('-date', '-created_at')[:50]
    # Summary by financial_type, category, description (year only)
    summary_table = TUFinancialTransaction.objects.filter(date__year=year)
    summary_table = summary_table.values('financial_type', 'category', 'description').annotate(total_amount=Sum('amount')).order_by('financial_type', 'category', 'description')
    # Convert category and description IDs to string for template
    for row in summary_table:
        cat_obj = FinancialCategory.objects.filter(pk=row['category']).first()
        desc_obj = FinancialDescription.objects.filter(pk=row['description']).first()
        row['category'] = str(cat_obj) if cat_obj else row['category']
        row['description'] = str(desc_obj) if desc_obj else row['description']
        row['estimated_expense'] = cat_obj.estimated_expense if cat_obj else 0
    form = TUFinancialTransactionForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        transaction = form.save(commit=False)
        transaction.created_by = Employee.objects.filter(user=request.user).first()
        transaction.save()
        messages.success(request, 'Financial transaction added!')
        return redirect('financial')
    # Category summary for all categories (show all, even if no transactions)
    categories_income = FinancialCategory.objects.filter(type='income').order_by('code')
    categories_expense = FinancialCategory.objects.filter(type='expense').order_by('code')
    total_amount_by_cat = TUFinancialTransaction.objects.filter(date__year=year).values('category').annotate(total=Sum('amount'))
    total_amount_map = {row['category']: row['total'] for row in total_amount_by_cat}
    # Build summary for all Category/Description pairs (Income/Outcome)
    income_summary = build_summary(year, 'income')
    outcome_summary = build_summary(year, 'expense')
    return render(request, 'employee/financial.html', {
        'form': form,
        'transactions': transactions,
        'tu_summary': tu_summary,
        'year': year,
        'month': month,
        'months': list(range(1, 13)),
        'is_superuser': is_superuser,
        'is_committee': is_committee,
        'summary_table': summary_table,
        'income_summary': income_summary,
        'outcome_summary': outcome_summary,
    })

# Edit financial transaction
@login_required
def edit_financial_transaction(request, pk):
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	if not (is_superuser or is_committee):
		return redirect('home')
	transaction = TUFinancialTransaction.objects.get(pk=pk)
	form = TUFinancialTransactionForm(request.POST or None, request.FILES or None, instance=transaction)
	if request.method == 'POST' and form.is_valid():
		transaction = form.save(commit=False)
		transaction.created_by = Employee.objects.filter(user=request.user).first()
		transaction.save()
		messages.success(request, 'Financial transaction updated!')
		return redirect('financial')
	return render(request, 'employee/financial.html', {
		'form': form,
		'transactions': TUFinancialTransaction.objects.order_by('-date', '-created_at')[:50],
		'edit_mode': True,
		'edit_id': pk,
		'is_superuser': is_superuser,
		'is_committee': is_committee,
	})

# Delete financial transaction
@login_required
def delete_financial_transaction(request, pk):
	is_superuser = request.user.is_superuser if request.user.is_authenticated else False
	is_committee = request.user.groups.filter(name='TU committee').exists() if request.user.is_authenticated else False
	if not (is_superuser or is_committee):
		return redirect('home')
	transaction = TUFinancialTransaction.objects.get(pk=pk)
	transaction.delete()
	messages.success(request, 'Financial transaction deleted!')
	return redirect('financial')

@login_required
def get_financial_options(request):
		ftype = request.GET.get('type')
		category_id = request.GET.get('category')
		# Nếu là club-financial thì chỉ lấy for_type club hoặc both
		is_club = 'club-financial' in request.META.get('HTTP_REFERER', '')
		if is_club:
			categories = FinancialCategory.objects.filter(type=ftype, for_type__in=['club', 'both'])
		else:
			categories = FinancialCategory.objects.filter(type=ftype)
		if category_id:
			descriptions = FinancialDescription.objects.filter(category_id=category_id)
		else:
			descriptions = FinancialDescription.objects.filter(type=ftype)
		return JsonResponse({
			'categories': [{'id': c.id, 'name': str(c), 'for_type': c.for_type} for c in categories],
			'descriptions': [{'id': d.id, 'text': str(d)} for d in descriptions],
		})

def get_tu_financial_summary(year, month=None):
    opening_obj = FinancialOpeningBalance.objects.filter(type='tu', year=year, month__isnull=True).first()
    opening_balance = opening_obj.opening_balance if opening_obj else 0
    transactions = TUFinancialTransaction.objects.filter(date__year=year)
    if month:
        transactions = transactions.filter(date__month=month)
    total_income = sum(t.amount for t in transactions if t.financial_type == 'income')
    total_expense = sum(t.amount for t in transactions if t.financial_type == 'expense')
    closing_balance = opening_balance + total_income - total_expense
    return {
        'year': year,
        'month': month,
        'opening_balance': opening_balance,
        'total_income': total_income,
        'total_expense': total_expense,
        'closing_balance': closing_balance,
    }

@login_required
def export_financial_report(request):
    year = int(request.GET.get('year', datetime.datetime.now().year))
    expenses = TUFinancialTransaction.objects.filter(date__year=year, financial_type='expense').order_by('date')
    # Path to your Excel template (update if needed)
    template_path = os.path.join(settings.BASE_DIR, 'employee', 'employee_import_template.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active
    # Write data starting from row 2 (assuming row 1 is header)
    row = 2
    for t in expenses:
        ws.cell(row=row, column=1, value=t.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=str(t.category))
        ws.cell(row=row, column=3, value=t.payment_id)
        ws.cell(row=row, column=4, value=t.description)
        ws.cell(row=row, column=5, value=t.details)
        ws.cell(row=row, column=6, value=t.amount)
        ws.cell(row=row, column=7, value=str(t.created_by) if t.created_by else '')
        row += 1
    filename = f'EndavaTU_BaoCaoThuChiYear{year}.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = FileResponse(output, as_attachment=True, filename=filename)
    return response

@login_required
def edit_club_financial_transaction(request, pk):
	if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
		return redirect('home')
	transaction = ClubFinancialTransaction.objects.get(pk=pk)
	form = ClubFinancialForm(request.POST or None, request.FILES or None, instance=transaction)
	if request.method == 'POST' and form.is_valid():
		transaction = form.save(commit=False)
		transaction.created_by = Employee.objects.filter(user=request.user).first()
		transaction.save()
		messages.success(request, 'Club financial transaction updated!')
		return redirect('club_financial')
	return render(request, 'employee/club_financial.html', {
		'form': form,
		'edit_mode': True,
		'edit_id': pk,
	})

@login_required
def delete_club_financial_transaction(request, pk):
	if not (request.user.is_superuser or request.user.groups.filter(name='TU committee').exists()):
		return redirect('home')
	transaction = ClubFinancialTransaction.objects.get(pk=pk)
	transaction.delete()
	messages.success(request, 'Club financial transaction deleted!')
	return redirect('club_financial')